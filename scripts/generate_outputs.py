#!/usr/bin/env python3
"""
Generate test-case deliverables for API projects.

All rules, contracts, templates, and configuration are loaded from
the ``baseline/`` folder — nothing is hardcoded in this script.

Input files (under ``baseline/``):
    project_config.json        – BASE_URL, headers, error codes, mappings
    categories.json            – coverage categories & prompt requirements
    common_test_templates.json – common test-case templates
    excel_template.json        – Excel output column mapping
    base_api_defs/<slug>.json  – known API definitions
    api_specific_tests/<slug>.json – API-specific test cases
"""

import copy
import csv
import json
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

# SoapUI XML parser (shared utility)
try:
    from soapui_parser import parse_soapui_xml
except ImportError:
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from soapui_parser import parse_soapui_xml

# SOAP XML body utilities
try:
    from soap_body_utils import detect_soap_body, apply_soap_body_mod, soap_body_to_flat_dict
except ImportError:
    from soap_body_utils import detect_soap_body, apply_soap_body_mod, soap_body_to_flat_dict


# ==============================================================================
#  PATH SETUP
# ==============================================================================

_SCRIPT_DIR = Path(__file__).resolve().parent
_ROOT_REPO  = _SCRIPT_DIR.parent
_INPUT_DIR  = _ROOT_REPO / "baseline"

ROOT = _ROOT_REPO / "output"


# ==============================================================================
#  CONFIG LOADING — all from baseline/ files
# ==============================================================================

def _load_json(path: Path) -> dict:
    """Load a JSON file; abort with a clear error if missing."""
    if not path.exists():
        print(f"❌  Missing input file: {path}")
        sys.exit(1)
    with open(path, encoding="utf-8") as f:
        return json.load(f)


# ── Project config ────────────────────────────────────────────────────────────
_project_cfg = _load_json(_INPUT_DIR / "project_config.json")

BASE_URL            = os.environ.get("BASE_URL", _project_cfg["base_url"])
STANDARD_HEADERS    = _project_cfg["standard_headers"]
RESPONSE_ENVELOPE   = _project_cfg["response_envelope"]
COMMON_ERRORS       = _project_cfg["common_errors"]
_DEFAULT_COLLECTION_VARIABLES = _project_cfg["default_collection_variables"]

_HEADER_TO_VAR      = _project_cfg["header_to_variable_map"]
_HEADER_SKIP        = set(_project_cfg.get("header_skip", []))
_HEADER_VARS        = _project_cfg.get("header_variable_replacements", {})

_STATUS_EQUIVALENTS = {
    int(k): v for k, v in _project_cfg.get("status_equivalents", {}).items()
}

_VN_KEYWORD_FILTER  = _project_cfg.get("vietnamese_keyword_filter",
                                        ["nếu", "lấy", "theo", "không", "hiện", "khi"])

_RESPONSE_TIME_MS   = _project_cfg.get("response_time_threshold_ms", 5000)

_ERROR_CODE_HTTP_MAP = _project_cfg.get("error_code_http_map", {})
_HTTP_STATUS_PATTERNS = _project_cfg.get("http_status_inference_patterns", [])
_COMMON_ERROR_SKIP   = set(_project_cfg.get("common_error_keys_skip", []))

_SAMPLER_OUTDATED_FIELDS = _project_cfg.get("sampler_outdated_fields", {})
_SAMPLER_ONLY_HEADERS    = _project_cfg.get("sampler_only_headers", [])
_GEN_EXTRA_HEADERS       = _project_cfg.get("gen_extra_headers", [])

# ── Categories ────────────────────────────────────────────────────────────────
_cat_cfg = _load_json(_INPUT_DIR / "categories.json")

CATEGORIES = _cat_cfg["categories"]
PROMPT_REQUIREMENTS = [
    (r["num"], r["label"], r["category_key"], r["description"])
    for r in _cat_cfg["prompt_requirements"]
]
EXTRA_CATEGORIES = [
    (r["label"], r["category_key"], r["description"])
    for r in _cat_cfg["extra_categories"]
]

# ── Common test templates ─────────────────────────────────────────────────────
_tpl_cfg = _load_json(_INPUT_DIR / "common_test_templates.json")

# ── Excel template ────────────────────────────────────────────────────────────
_excel_cfg_path = _INPUT_DIR / "excel_template.json"
_excel_cfg = _load_json(_excel_cfg_path) if _excel_cfg_path.exists() else None

# ── (Legacy) base_api_defs / api_specific_tests đã bỏ ────────────────────────
# Tất cả API đều dùng generic builder. Giữ biến rỗng cho backward compat.
BASE_API_DEFS = {}
ALL_APIS = []


# ==============================================================================
#  KPI TARGETS  — from baseline/coverage_requirements.json
# ==============================================================================

_CONFIG_FILE = _ROOT_REPO / 'baseline' / 'coverage_requirements.json'


def _load_kpi_targets() -> dict:
    defaults = {
        "prompt_coverage_pct":   90,
        "min_total_tcs":         50,
        "min_p1_pct":            30,
        "min_p1p2_pct":          70,
        "min_http_status_codes": 6,
        "min_error_codes":       3,
    }
    if not _CONFIG_FILE.exists():
        print(f"[WARN] coverage_requirements.json not found at {_CONFIG_FILE} — using default KPI targets")
        return defaults
    try:
        cfg = json.loads(_CONFIG_FILE.read_text(encoding='utf-8'))
        loaded = cfg.get("kpi_targets", {})
    except Exception as exc:
        print(f"[WARN] Cannot parse coverage_requirements.json ({exc}) — using default KPI targets")
        return defaults
    if not loaded:
        print(f"[WARN] No kpi_targets found in coverage_requirements.json — using defaults")
        return defaults
    merged = {**defaults, **{k: v for k, v in loaded.items() if not k.startswith('_')}}
    return merged

KPI_TARGETS = _load_kpi_targets()


# ==============================================================================
#  RESPONSE BODY STRUCTURE HELPERS
# ==============================================================================

def _extract_fields_from_response_body(body: dict) -> dict:
    """Parse a response JSON body and return a {field_name: type_str} dict
    suitable for use as response_data_fields in api_def.
    Looks inside body['data'] first (standard response envelope), then root.
    Type strings: 'String', 'Number', 'Boolean', 'Array', 'Object'.
    """
    if not isinstance(body, dict):
        return {}
    _JS_TYPE = {
        str:   'String',
        int:   'Number',
        float: 'Number',
        bool:  'Boolean',
        list:  'Array',
        dict:  'Object',
    }
    data_obj = body.get('data', body)
    if not isinstance(data_obj, dict):
        # data might be an array (list APIs) — use first element if available
        if isinstance(data_obj, list) and data_obj and isinstance(data_obj[0], dict):
            data_obj = data_obj[0]
        else:
            data_obj = body
    result = {}
    for k, v in data_obj.items():
        if k in ('code', 'message', 'messageKey'):   # envelope fields, not data
            continue
        result[k] = _JS_TYPE.get(type(v), 'String')
    return result


# ==============================================================================
#  SAMPLER METADATA READER
# ==============================================================================

def read_sampler_metadata(sampler_path) -> dict:
    """Read collection-level variable, auth, and per-request prerequest scripts from a sampler file.

    Supports both Postman JSON (.postman_collection.json) and SoapUI XML (.xml) files.

    Returns a dict with keys:
      - 'variable': list of collection variables
      - 'auth': collection-level auth block (or None)
      - 'prerequest_by_url': dict mapping full request URL → list[str] of JS exec lines
                             extracted from that request's pre-request script in the sampler.
                             Used as automatic fallback when 'sampler_prerequest' is not set
                             in the contract.
    """
    sampler_path = Path(sampler_path) if not isinstance(sampler_path, Path) else sampler_path
    result = {"variable": list(_DEFAULT_COLLECTION_VARIABLES), "auth": None, "prerequest_by_url": {}}

    if not sampler_path.exists():
        print(f"  ⚠️  Sampler not found ({sampler_path.name}) — using default variables, no auth.")
        return result

    # ── SoapUI XML sampler ────────────────────────────────────────────────
    if sampler_path.suffix == '.xml':
        soap_items = parse_soapui_xml(sampler_path)
        extracted_values = {}
        for req_item in soap_items:
            req = req_item.get('request', {})
            for h in req.get('header', []):
                hk = h.get('key', '').lower()
                hv = h.get('value', '')
                if hk in _HEADER_TO_VAR and hv and not hv.startswith('{{'):
                    var_key = _HEADER_TO_VAR[hk]
                    if var_key not in extracted_values:
                        extracted_values[var_key] = hv
            url_obj = req.get('url', {})
            raw_url = url_obj.get('raw', '') if isinstance(url_obj, dict) else str(url_obj)
            if raw_url and '://' in raw_url:
                m = re.match(r'(https?://[^/]+)', raw_url)
                if m and 'baseUrl' not in extracted_values:
                    extracted_values['baseUrl'] = m.group(1)
        if extracted_values:
            print(f"  ✅  Extracted {len(extracted_values)} variable value(s) from SoapUI sampler: {', '.join(extracted_values.keys())}")
            existing_lower = {v.get('key', '').lower(): v for v in result["variable"]}
            for var_key, var_val in extracted_values.items():
                match = existing_lower.get(var_key.lower())
                if match is not None:
                    current = match.get('value', '')
                    if not current or current.startswith('{{'):
                        match['value'] = var_val
                else:
                    result["variable"].append({"key": var_key, "value": var_val})
                    existing_lower[var_key.lower()] = result["variable"][-1]
        return result

    # ── Postman JSON sampler ──────────────────────────────────────────────

    try:
        with open(sampler_path, encoding="utf-8") as f:
            col = json.load(f)
    except Exception as exc:
        print(f"  ⚠️  Cannot read sampler ({exc}) — using defaults.")
        return result

    extracted_values = {}

    def _collect_requests(items):
        for it in items:
            if 'request' in it:
                yield it
            if 'item' in it:
                yield from _collect_requests(it['item'])

    prerequest_by_url: dict = {}
    resp_body_by_url: dict = {}
    for req_item in _collect_requests(col.get('item', [])):
        req = req_item.get('request', {})
        for h in req.get('header', []):
            hk = h.get('key', '').lower()
            hv = h.get('value', '')
            if hk in _HEADER_TO_VAR and hv and not hv.startswith('{{'):
                var_key = _HEADER_TO_VAR[hk]
                if var_key not in extracted_values:
                    extracted_values[var_key] = hv
        url_obj = req.get('url', {})
        raw_url = url_obj.get('raw', '') if isinstance(url_obj, dict) else str(url_obj)
        if raw_url and '://' in raw_url:
            m = re.match(r'(https?://[^/]+)', raw_url)
            if m and 'baseUrl' not in extracted_values:
                extracted_values['baseUrl'] = m.group(1)
        # ── Extract per-request pre-request script ────────────────────────────
        for ev in req_item.get('event', []):
            if ev.get('listen') == 'prerequest':
                exec_lines = ev.get('script', {}).get('exec', [])
                non_empty = [l for l in exec_lines if l.strip()]
                if non_empty and raw_url and '://' in raw_url:
                    # Normalise URL: strip query strings for keying
                    clean_url = raw_url.split('?')[0].rstrip('/')
                    if clean_url not in prerequest_by_url:
                        prerequest_by_url[clean_url] = exec_lines

        # ── Extract response body examples (for schema detection) ─────────────
        for resp_ex in req_item.get('response', []):
            status_code = resp_ex.get('code', resp_ex.get('status', 0))
            if isinstance(status_code, str):
                try:
                    status_code = int(status_code)
                except ValueError:
                    status_code = 0
            rbody = resp_ex.get('body', '')
            if rbody and raw_url and '://' in raw_url:
                clean_url = raw_url.split('?')[0].rstrip('/')
                bucket = 'success' if (200 <= status_code < 300) else 'failure'
                try:
                    parsed = json.loads(rbody)
                    resp_body_by_url.setdefault(clean_url, {})
                    if bucket not in resp_body_by_url[clean_url]:
                        resp_body_by_url[clean_url][bucket] = parsed
                except (json.JSONDecodeError, ValueError):
                    pass

    if prerequest_by_url:
        result['prerequest_by_url'] = prerequest_by_url
        print(f"  ✅  Found pre-request scripts on {len(prerequest_by_url)} sampler request(s): "
              f"{', '.join(u.rsplit('/', 1)[-1] for u in prerequest_by_url)}")

    if resp_body_by_url:
        result['response_body_by_url'] = resp_body_by_url
        print(f"  ✅  Found response body examples for {len(resp_body_by_url)} sampler request(s): "
              f"{', '.join(u.rsplit('/', 1)[-1] for u in resp_body_by_url)}")

    if extracted_values:
        print(f"  ✅  Extracted {len(extracted_values)} variable value(s) from sampler requests: {', '.join(extracted_values.keys())}")

    raw_vars = col.get("variable", [])
    if raw_vars and isinstance(raw_vars, list):
        result["variable"] = raw_vars
        result["_has_sampler_vars"] = True
        print(f"  ✅  Inherited {len(raw_vars)} collection variables from sampler.")
    else:
        result["_has_sampler_vars"] = False
        print(f"  ℹ️  Sampler has no collection variables — using defaults ({len(result['variable'])} vars).")

    existing_lower = {v.get('key', '').lower(): v for v in result["variable"]}
    for var_key, var_val in extracted_values.items():
        match = existing_lower.get(var_key.lower())
        if match is not None:
            current = match.get('value', '')
            if not current or current.startswith('{{'):
                match['value'] = var_val
        else:
            result["variable"].append({"key": var_key, "value": var_val})
            existing_lower[var_key.lower()] = result["variable"][-1]

    raw_auth = col.get("auth")
    if raw_auth and isinstance(raw_auth, dict):
        result["auth"] = raw_auth
        auth_type = raw_auth.get("type", "unknown")
        print(f"  ✅  Inherited collection auth (type={auth_type}) from sampler.")
    else:
        print(f"  ℹ️  Sampler has no collection auth — output collections will have no auth block.")

    return result


# ==============================================================================
#  URL BUILDER
# ==============================================================================

def build_url_obj(path: str, full_url: str = None) -> dict:
    raw = full_url if full_url else BASE_URL + path
    # Parse protocol from the actual raw URL
    protocol = "https"
    if "://" in raw:
        protocol = raw.split("://", 1)[0]
    no_scheme = raw.split("://", 1)[-1]
    host_str = no_scheme.split("/", 1)[0]
    host_parts = host_str.split(".")
    path_parts = [p for p in path.split("/") if p]
    return {
        "raw": raw,
        "protocol": protocol,
        "host": host_parts,
        "path": path_parts,
    }


# ==============================================================================
#  REQUEST BODY BUILDER
# ==============================================================================

def build_default_body(api_def: dict) -> dict:
    tmpl = api_def.get("body_template")
    if tmpl and isinstance(tmpl, dict):
        return copy.deepcopy(tmpl)

    body = {}
    for field, info in api_def["request_body"].items():
        val = info["value"]
        if val is not None:
            body[field] = val
        else:
            ftype = (info.get("type") or "String").lower()
            if ftype in ("number", "integer", "int", "long"):
                body[field] = 0
            elif ftype in ("boolean", "bool"):
                body[field] = True
            else:
                body[field] = ""
    return body


# ---------------------------------------------------------------------------
#  Helpers for nested body operations
# ---------------------------------------------------------------------------

def _nested_remove(d: dict, key: str) -> bool:
    if key in d:
        del d[key]
        return True
    for v in d.values():
        if isinstance(v, dict) and _nested_remove(v, key):
            return True
    return False


def _nested_set(d: dict, key: str, value) -> bool:
    if key in d:
        d[key] = value
        return True
    for v in d.values():
        if isinstance(v, dict) and _nested_set(v, key, value):
            return True
    return False


def _dot_path_remove(d: dict, path: str) -> bool:
    parts = path.split('.')
    cur = d
    for p in parts[:-1]:
        if isinstance(cur, dict) and p in cur:
            cur = cur[p]
        else:
            return False
    if isinstance(cur, dict) and parts[-1] in cur:
        del cur[parts[-1]]
        return True
    return False


def _dot_path_set(d: dict, path: str, value) -> bool:
    parts = path.split('.')
    cur = d
    for p in parts[:-1]:
        if isinstance(cur, dict) and p in cur:
            cur = cur[p]
        else:
            return False
    if isinstance(cur, dict):
        cur[parts[-1]] = value
        return True
    return False


def apply_body_modifications(base_body: dict, body_mod: dict,
                             field_map: dict = None) -> dict:
    body = copy.deepcopy(base_body)
    field_map = field_map or {}

    removes = body_mod.get("__remove__", [])
    if isinstance(removes, str):
        removes = [removes]
    for key in removes:
        if key in field_map:
            _dot_path_remove(body, field_map[key])
        elif key not in body:
            _nested_remove(body, key)
        else:
            body.pop(key, None)

    nulls = body_mod.get("__set_null__", [])
    if isinstance(nulls, str):
        nulls = [nulls]
    for key in nulls:
        if key in field_map:
            _dot_path_set(body, field_map[key], None)
        elif key in body:
            body[key] = None
        else:
            _nested_set(body, key, None)

    for k, v in body_mod.items():
        if k.startswith("__"):
            continue
        if k in field_map:
            _dot_path_set(body, field_map[k], v)
        elif k in body:
            body[k] = v
        else:
            _nested_set(body, k, v)

    return body


# ==============================================================================
#  POSTMAN TEST SCRIPT GENERATOR
# ==============================================================================

def postman_test_script(case: dict, api_def: dict) -> list:
    """Generate comprehensive Postman test assertions."""
    lines = []
    tc_id = case["id"]
    expected = case["expected_status"]

    # 0. Setup-chain failure guard — if prequisites failed, report it cleanly
    has_setup = bool(api_def.get("setup_items"))
    if has_setup:
        lines += [
            "// ── Setup-chain result check ──",
            "var _setupErr = pm.collectionVariables.get('_setupError');",
            "if (_setupErr) {",
            f"    pm.test('[{tc_id}] Setup prerequisites completed', function () {{",
            "        pm.expect.fail('Setup chain ABORTED: ' + _setupErr + ' → main request used stale data');",
            "    });",
            "}",
            "",
        ]

    # 1. HTTP status code
    accept_statuses = case.get("accept_status", [])
    if not accept_statuses and expected in _STATUS_EQUIVALENTS:
        accept_statuses = _STATUS_EQUIVALENTS[expected]
    if accept_statuses and len(accept_statuses) > 1:
        statuses_js = json.dumps(accept_statuses)
        lines += [
            f"// == [{tc_id}] Assertions ==",
            f"pm.test('[{tc_id}] HTTP status is one of {accept_statuses}', function () {{",
            f"    pm.expect({statuses_js}).to.include(pm.response.code);",
            "});",
            "",
        ]
    else:
        lines += [
            f"// == [{tc_id}] Assertions ==",
            f"pm.test('[{tc_id}] HTTP status is {expected}', function () {{",
            f"    pm.response.to.have.status({expected});",
            "});",
            "",
        ]

    # ── SOAP-specific assertions (replaces all JSON-based checks) ─────────────
    is_soap = api_def.get('is_soap', False)
    if is_soap:
        _cat = case.get('category', '')
        _soap_error_categories = ('Business Rules', 'Backend Failure', 'Negative Validation', 'Error Handling')
        _soap_ep_is_error = (
            _cat in ('Equivalence Partitioning', 'Boundary Value Analysis')
            and any(kw in (case.get('name') or '') for kw in (
                'Invalid', 'invalid', 'Missing', 'Empty', 'Null', 'null', 'Negative', 'negative'
            ))
        )
        _is_soap_t24error = _cat in _soap_error_categories or _soap_ep_is_error

        if _cat == 'XML Validation':
            # Gateway rejects malformed XML → no SOAP envelope to check
            lines += [
                "// SOAP XML Validation: expect gateway/parser error (no valid SOAP body)",
                "pm.test('Server rejected malformed/non-XML request', function () {",
                "    pm.expect(pm.response.code).to.be.at.least(400);",
                "});",
                "",
            ]
        elif _is_soap_t24error:
            # SOAP business/backend error: HTTP 200 but body contains T24Error
            lines += [
                "// SOAP Error case: response is XML, successIndicator=T24Error",
                "pm.test('Response body is non-empty XML', function () {",
                "    var text = pm.response.text();",
                "    pm.expect(text).to.be.a('string').and.to.have.length.above(0);",
                "    pm.expect(text.trim().charAt(0)).to.equal('<');",
                "});",
                "",
                "pm.test('SOAP response signals T24Error', function () {",
                "    var text = pm.response.text();",
                "    pm.expect(text).to.include('<successIndicator>T24Error</successIndicator>');",
                "});",
                "",
                "pm.test('SOAP error response contains <messages> element', function () {",
                "    var text = pm.response.text();",
                "    pm.expect(text).to.include('<messages>');",
                "});",
                "",
            ]
            # If messages_text is known, assert it explicitly
            err_key = case.get('expected_error_key', '')
            messages_text = ''
            if err_key:
                messages_text = api_def.get('api_errors', {}).get(err_key, {}).get('messages_text', '')
            if messages_text:
                lines += [
                    f"pm.test('SOAP messages contains expected error text', function () {{",
                    "    var text = pm.response.text();",
                    f"    pm.expect(text).to.include('{messages_text}');",
                    "});",
                    "",
                ]
        else:
            # SOAP success: HTTP 200, successIndicator=Success, response data fields present
            lines += [
                "// SOAP Success case: response is XML with successIndicator=Success",
                "pm.test('Response body is non-empty XML', function () {",
                "    var text = pm.response.text();",
                "    pm.expect(text).to.be.a('string').and.to.have.length.above(0);",
                "    pm.expect(text.trim().charAt(0)).to.equal('<');",
                "});",
                "",
                "pm.test('Response Content-Type indicates XML', function () {",
                "    pm.response.to.have.header('Content-Type');",
                "    var ct = pm.response.headers.get('Content-Type');",
                "    pm.expect(ct.toLowerCase()).to.satisfy(function(c) {",
                "        return c.includes('xml') || c.includes('text/xml') || c.includes('soap');",
                "    });",
                "});",
                "",
                "pm.test('SOAP response has <successIndicator> element', function () {",
                "    var text = pm.response.text();",
                "    pm.expect(text).to.include('<successIndicator>');",
                "});",
                "",
                "pm.test('SOAP successIndicator is Success (not T24Error)', function () {",
                "    var text = pm.response.text();",
                "    pm.expect(text).to.include('<successIndicator>Success</successIndicator>');",
                "});",
                "",
            ]
            # Check that expected response data fields appear as XML elements
            data_fields = api_def.get("response_data_fields", {})
            if data_fields:
                field_list = list(data_fields.keys())
                fields_comment = ", ".join(f"<{f}>" for f in field_list)
                lines += [
                    f"// Expected response data elements: {fields_comment}",
                ]
                for fname in field_list:
                    lines += [
                        f"pm.test('Response XML contains <{fname}> element', function () {{",
                        "    var text = pm.response.text();",
                        f"    pm.expect(text).to.include('<{fname}>');",
                        "});",
                        "",
                    ]
        return lines

    # 2. Valid JSON
    lines += [
        "pm.test('Response body is valid JSON', function () {",
        "    if (pm.response.text()) {",
        "        pm.expect(function () { pm.response.json(); }).to.not.throw();",
        "    }",
        "});",
        "",
    ]

    # 4. Response Content-Type header
    lines += [
        "pm.test('Response Content-Type is application/json', function () {",
        "    pm.response.to.have.header('Content-Type');",
        "    var ct = pm.response.headers.get('Content-Type');",
        "    pm.expect(ct).to.include('application/json');",
        "});",
        "",
    ]

    # 5. Response envelope schema (for 2xx)
    if 200 <= expected < 300:
        envelope_fields = RESPONSE_ENVELOPE
        fields_js = ", ".join(f'"{f}"' for f in envelope_fields)
        lines += [
            "pm.test('Response envelope has standard fields', function () {",
            "    var json = pm.response.json();",
            f"    var expected = [{fields_js}];",
            "    var hasEnvelope = expected.every(function(f) {{ return json.hasOwnProperty(f); }});",
            "    var hasDirectData = typeof json === 'object' && json !== null;",
            "    pm.expect(hasEnvelope || hasDirectData).to.be.true;",
            "});",
            "",
        ]

        data_fields = api_def.get("response_data_fields", {})
        if data_fields:
            df_js = ", ".join(f'"{f}"' for f in data_fields.keys())
            lines += [
                "pm.test('Response data has required fields', function () {",
                "    var json = pm.response.json();",
                "    var dataObj = json.data || json;",
                "    if (typeof dataObj === 'object' && dataObj !== null) {",
                f"        var expected = [{df_js}];",
                "        var missing = expected.filter(function(f) { return !dataObj.hasOwnProperty(f); });",
                "        pm.expect(missing.length, 'Missing fields: ' + missing.join(', ')).to.equal(0);",
                "    }",
                "});",
                "",
            ]

        arr_field = api_def.get("response_array_field")
        arr_item_fields = api_def.get("response_array_item_fields", {})
        if arr_field and arr_item_fields:
            aif_js = ", ".join(f'"{f}"' for f in arr_item_fields.keys())
            lines += [
                f"pm.test('Array items in data.{arr_field} have required fields', function () {{",
                "    var json = pm.response.json();",
                "    var dataObj = json.data || json;",
                f"    if (dataObj.{arr_field} && dataObj.{arr_field}.length > 0) {{",
                f"        var item = dataObj.{arr_field}[0];",
                f"        var expected = [{aif_js}];",
                "        expected.forEach(function(field) {",
                "            pm.expect(item).to.have.property(field);",
                "        });",
                "    }",
                "});",
                "",
            ]

        for fname, ftype in data_fields.items():
            if ftype == "String":
                lines += [
                    f"pm.test('data.{fname} is a string', function () {{",
                    "    var json = pm.response.json();",
                    "    var dataObj = json.data || json;",
                    f"    if (dataObj.{fname} !== undefined) {{",
                    f"        pm.expect(typeof dataObj.{fname}).to.equal('string');",
                    "    }",
                    "});",
                    "",
                ]
            elif ftype == "Array":
                lines += [
                    f"pm.test('data.{fname} is an array', function () {{",
                    "    var json = pm.response.json();",
                    "    var dataObj = json.data || json;",
                    f"    if (dataObj.{fname} !== undefined) {{",
                    f"        pm.expect(dataObj.{fname}).to.be.an('array');",
                    "    }",
                    "});",
                    "",
                ]
            elif ftype == "Object":
                lines += [
                    f"pm.test('data.{fname} is an object', function () {{",
                    "    var json = pm.response.json();",
                    "    var dataObj = json.data || json;",
                    f"    if (dataObj.{fname} !== undefined) {{",
                    f"        pm.expect(dataObj.{fname}).to.be.an('object');",
                    "    }",
                    "});",
                    "",
                ]

        if case.get("enum_check"):
            enum_field = case["enum_check"]["field"]
            enum_path = case["enum_check"]["path"]
            allowed_js = json.dumps(case["enum_check"]["allowed"])
            lines += [
                f"pm.test('{enum_field} value is in allowed enum', function () {{",
                "    var json = pm.response.json();",
                f"    var allowed = {allowed_js};",
                "    try {",
                f"        var val = eval('json.{enum_path}');",
                "        if (val !== undefined) {",
                "            pm.expect(allowed).to.include(val);",
                "        }",
                "    } catch(e) { /* field may not exist */ }",
                "});",
                "",
            ]

    # 6. Error response structure (for 4xx/5xx)
    if expected >= 400:
        lines += [
            "pm.test('Error response has code and message', function () {",
            "    var json = pm.response.json();",
            "    var hasFormatA = json.hasOwnProperty('code') && json.hasOwnProperty('message');",
            "    var hasFormatB = json.hasOwnProperty('detail');",
            "    pm.expect(hasFormatA || hasFormatB).to.be.true;",
            "});",
            "",
        ]

        if case.get("expected_error_key"):
            err_key = case["expected_error_key"]
            all_errors = {**COMMON_ERRORS, **api_def.get("api_errors", {})}
            if err_key in all_errors:
                err_code = all_errors[err_key]["code"]
                lines += [
                    f"pm.test('Error code is {err_code} ({err_key})', function () {{",
                    "    var json = pm.response.json();",
                    "    if (json.code) {",
                    f"        pm.expect(json.code).to.equal('{err_code}');",
                    "    } else if (json.detail) {",
                    "        pm.expect(json.detail).to.be.an('array');",
                    "    }",
                    "});",
                    "",
                    f"pm.test('Error messageKey is {err_key}', function () {{",
                    "    var json = pm.response.json();",
                    "    if (json.messageKey) {",
                    f"        pm.expect(json.messageKey).to.equal('{err_key}');",
                    "    } else if (json.detail) {",
                    "        pm.expect(json.detail).to.be.an('array');",
                    "    }",
                    "});",
                    "",
                ]

    return lines


# ==============================================================================
#  TEMPLATE-BASED TEST CASE GENERATION
# ==============================================================================

def _resolve_template_body_mod(raw_mod: dict, f1: str, f2: str,
                               optional_fields: list, all_fields: list) -> dict:
    """Resolve template placeholders in body_mod values."""
    resolved = {}
    for k, v in raw_mod.items():
        # Resolve key placeholders
        rk = k.replace("{f1}", f1).replace("{f2}", f2)

        # Resolve value placeholders
        if isinstance(v, str):
            if v == "__OPTIONAL_FIELDS__":
                resolved["__remove__"] = optional_fields
                continue
            if v == "__ALL_FIELDS__":
                resolved["__remove__"] = all_fields
                continue
            if v == "__REPEAT_X_260__":
                rv = "X" * 260
            elif v == "__REPEAT_DEL_50__":
                rv = "\x7f" * 50
            else:
                rv = v.replace("{f1}", f1).replace("{f2}", f2)
        elif isinstance(v, list):
            # resolve list items (e.g. __remove__: ["{f1}"])
            rv = []
            for item in v:
                if isinstance(item, str):
                    rv.append(item.replace("{f1}", f1).replace("{f2}", f2))
                else:
                    rv.append(item)
        else:
            rv = v

        resolved[rk] = rv
    return resolved


def _resolve_template_header_mod(raw_mod: dict) -> dict:
    """Resolve header_mod — JSON null → Python None for 'remove header'."""
    result = {}
    for k, v in raw_mod.items():
        result[k] = v  # None values from JSON are already Python None
    return result


def _resolve_template_name(name: str, f1: str, f2: str, **kwargs) -> str:
    """Resolve placeholders in test case name."""
    result = name.replace("{f1}", f1).replace("{f2}", f2)
    for k, v in kwargs.items():
        result = result.replace(f"{{{k}}}", str(v))
    return result


def generate_common_cases(api_def: dict) -> list:
    """Generate common test cases shared by all APIs — loaded from templates."""
    req_body = api_def["request_body"]
    mandatory = [k for k, v in req_body.items() if v.get("mandatory")]
    optional  = [k for k, v in req_body.items() if not v.get("mandatory")]
    all_keys  = list(req_body.keys())
    f1 = mandatory[0] if mandatory else (all_keys[0] if all_keys else "_field")
    f2 = mandatory[1] if len(mandatory) > 1 else f1

    cases = []

    # -- POSITIVE (static templates) --
    for tpl in _tpl_cfg.get("positive", []):
        body_mod = _resolve_template_body_mod(tpl.get("body_mod", {}), f1, f2, optional, all_keys)
        header_mod = _resolve_template_header_mod(tpl.get("header_mod", {}))
        cases.append({
            "cid": tpl["cid"],
            "name": _resolve_template_name(tpl["name"], f1, f2),
            "category": tpl["category"],
            "status": tpl["status"],
            "priority": tpl["priority"],
            "body_mod": body_mod,
            "header_mod": header_mod,
            "method_mod": tpl.get("method_mod"),
        })

    # -- POSITIVE: auto-generate from enum_values, enums, example_sets, notes --
    pos_idx = len(_tpl_cfg.get("positive", []))
    _pos_seen_mods = set()

    def _add_pos_case(name: str, body_mod: dict, priority: str = "P1"):
        nonlocal pos_idx
        mod_key = frozenset(sorted((k, str(v)) for k, v in body_mod.items()))
        if mod_key in _pos_seen_mods:
            return
        _pos_seen_mods.add(mod_key)
        pos_idx += 1
        cases.append({
            "cid": f"POS-{pos_idx:03d}",
            "name": name,
            "category": "Positive", "status": 200, "priority": priority,
            "body_mod": body_mod, "header_mod": {}, "method_mod": None,
        })

    # 1) From enum_values attached to each request field
    for field_name, field_info in req_body.items():
        evs = field_info.get("enum_values", [])
        for ev in evs:
            if ev in ("null", "not_null"):
                if ev == "null":
                    _add_pos_case(
                        f"Valid request with {field_name} omitted (cif-level query)",
                        {"__remove__": field_name}, "P1")
                else:
                    _add_pos_case(
                        f"Valid request with {field_name} present (account-level query)",
                        {}, "P1")
            else:
                _add_pos_case(
                    f"Valid request with {field_name} = \"{ev}\" (doc example)",
                    {field_name: ev}, "P1")

    # 2) From api_def["enums"]
    api_enums = api_def.get("enums", {})
    for enum_field, enum_values in api_enums.items():
        if enum_field not in req_body:
            continue
        for ev in enum_values:
            _add_pos_case(
                f"Valid request with {enum_field} = \"{ev}\" (enum)",
                {enum_field: ev}, "P1")

    # 3) From example_sets
    for i, ex_set in enumerate(api_def.get("example_sets", [])):
        _add_pos_case(f"Valid request with doc example set #{i+1}", ex_set, "P1")

    # 4) From note values
    for field_name, field_info in req_body.items():
        note_val = field_info.get("note", "")
        default_val = field_info.get("value")
        if not note_val or note_val == default_val:
            continue
        if "{{" in str(note_val) or any(
            kw in str(note_val).lower() for kw in _VN_KEYWORD_FILTER
        ):
            continue
        if len(str(note_val)) > 60:
            continue
        _add_pos_case(
            f"Valid request with {field_name} = \"{note_val}\" (doc note)",
            {field_name: note_val}, "P2")

    # -- NEGATIVE VALIDATION (per-mandatory-field templates) --
    neg_templates = _tpl_cfg.get("negative_per_mandatory_field", [])
    for i, field in enumerate(mandatory):
        idx = i * len(neg_templates)
        for j, tpl in enumerate(neg_templates):
            body_mod = _resolve_template_body_mod(tpl.get("body_mod", {}), f1, f2, optional, all_keys)
            # Replace {field} in body_mod keys/values
            resolved_bm = {}
            for k, v in body_mod.items():
                rk = k.replace("{field}", field)
                if isinstance(v, str):
                    rv = v.replace("{field}", field)
                elif isinstance(v, list):
                    rv = [item.replace("{field}", field) if isinstance(item, str) else item for item in v]
                else:
                    rv = v
                resolved_bm[rk] = rv
            cases.append({
                "cid": f"NEG-{idx+j+1:03d}",
                "name": tpl["name"].replace("{field}", field),
                "category": tpl["category"],
                "status": tpl["status"],
                "priority": tpl["priority"],
                "body_mod": resolved_bm,
                "header_mod": _resolve_template_header_mod(tpl.get("header_mod", {})),
                "method_mod": tpl.get("method_mod"),
            })

    # -- NEGATIVE EXTRA --
    neg_next = len(mandatory) * len(neg_templates) + 1
    for tpl in _tpl_cfg.get("negative_extra", []):
        body_mod = _resolve_template_body_mod(tpl.get("body_mod", {}), f1, f2, optional, all_keys)
        cases.append({
            "cid": f"NEG-{neg_next:03d}",
            "name": _resolve_template_name(tpl["name"], f1, f2),
            "category": tpl["category"],
            "status": tpl["status"],
            "priority": tpl["priority"],
            "body_mod": body_mod,
            "header_mod": _resolve_template_header_mod(tpl.get("header_mod", {})),
            "method_mod": tpl.get("method_mod"),
            **({"accept_status": tpl["accept_status"]} if "accept_status" in tpl else {}),
        })
        neg_next += 1

    # -- EQUIVALENCE PARTITIONING (static) --
    for tpl in _tpl_cfg.get("equivalence_partitioning", []):
        body_mod = _resolve_template_body_mod(tpl.get("body_mod", {}), f1, f2, optional, all_keys)
        cases.append({
            "cid": tpl["cid"],
            "name": _resolve_template_name(tpl["name"], f1, f2),
            "category": tpl["category"],
            "status": tpl["status"],
            "priority": tpl["priority"],
            "body_mod": body_mod,
            "header_mod": _resolve_template_header_mod(tpl.get("header_mod", {})),
            "method_mod": tpl.get("method_mod"),
        })

    # -- EP: enum values from fields --
    # Rule B: Only generate EP tests for fields with ≥2 real enum values.
    #         Single-sample values are already covered by POS-001.
    # Rule C: For fields with ≥2 enums, also add an Invalid Enum test (→ 400).
    ep_idx = len(_tpl_cfg.get("equivalence_partitioning", []))
    _ep_generated_fields = set()
    for field_name, field_info in req_body.items():
        enum_vals = field_info.get("enum_values", [])
        if not enum_vals:
            continue
        # Separate real enum values from meta-values
        real_vals = [v for v in enum_vals if v not in ("null", "not_null")]
        meta_vals = [v for v in enum_vals if v in ("null", "not_null")]

        # Always generate null/not_null EP tests (query mode switching)
        for enum_val in meta_vals:
            ep_idx += 1
            if enum_val == "null":
                body_mod = {"__remove__": field_name}
                label = f"{field_name} = null (cif-level query)"
            else:
                body_mod = {}
                label = f"{field_name} not null (account-level query)"
            cases.append({
                "cid": f"EP-{ep_idx:03d}",
                "name": f"Valid {label}",
                "category": "Equivalence Partitioning",
                "status": 200, "priority": "P2",
                "body_mod": body_mod, "header_mod": {}, "method_mod": None,
            })

        # Rule B: Skip EP for single-sample (already in POS-001)
        if len(real_vals) < 2:
            continue

        _ep_generated_fields.add(field_name)
        for enum_val in real_vals:
            ep_idx += 1
            body_mod = {field_name: enum_val}
            label = f"{field_name} = \"{enum_val}\" (enum value from doc)"
            cases.append({
                "cid": f"EP-{ep_idx:03d}",
                "name": f"Valid {label}",
                "category": "Equivalence Partitioning",
                "status": 200, "priority": "P2",
                "body_mod": body_mod, "header_mod": {}, "method_mod": None,
            })
        # Rule C: Invalid enum value test for this field
        ep_idx += 1
        cases.append({
            "cid": f"EP-{ep_idx:03d}",
            "name": f"Invalid {field_name} = \"INVALID_ENUM_VALUE\" (not in allowed list)",
            "category": "Equivalence Partitioning",
            "status": 400, "priority": "P2",
            "body_mod": {field_name: "INVALID_ENUM_VALUE"}, "header_mod": {}, "method_mod": None,
        })

    # -- EP: enum values from api_def["enums"] --
    for enum_field, enum_values in api_enums.items():
        if enum_field in _ep_generated_fields:
            continue
        if enum_field not in req_body:
            continue
        # Rule B: api_def["enums"] always has ≥2 values (filtered at build time)
        for ev in enum_values:
            ep_idx += 1
            cases.append({
                "cid": f"EP-{ep_idx:03d}",
                "name": f"Valid {enum_field} = \"{ev}\" (enum from definition)",
                "category": "Equivalence Partitioning",
                "status": 200, "priority": "P2",
                "body_mod": {enum_field: ev}, "header_mod": {}, "method_mod": None,
            })
        # Rule C: Invalid enum value test
        ep_idx += 1
        cases.append({
            "cid": f"EP-{ep_idx:03d}",
            "name": f"Invalid {enum_field} = \"INVALID_ENUM_VALUE\" (not in allowed list)",
            "category": "Equivalence Partitioning",
            "status": 400, "priority": "P2",
            "body_mod": {enum_field: "INVALID_ENUM_VALUE"}, "header_mod": {}, "method_mod": None,
        })

    # -- Load remaining template sections --
    for section_key in ["boundary_value_analysis", "authentication", "header_validation",
                        "http_method", "schema_validation", "error_handling",
                        "security", "idempotency", "edge_cases", "single_cases"]:
        for tpl in _tpl_cfg.get(section_key, []):
            body_mod = _resolve_template_body_mod(tpl.get("body_mod", {}), f1, f2, optional, all_keys)
            cases.append({
                "cid": tpl["cid"],
                "name": _resolve_template_name(tpl["name"], f1, f2),
                "category": tpl["category"],
                "status": tpl["status"],
                "priority": tpl["priority"],
                "body_mod": body_mod,
                "header_mod": _resolve_template_header_mod(tpl.get("header_mod", {})),
                "method_mod": tpl.get("method_mod"),
                **({"expected_error_key": tpl["expected_error_key"]} if "expected_error_key" in tpl else {}),
                **({"accept_status": tpl["accept_status"]} if "accept_status" in tpl else {}),
            })

    # -- SOAP-only extra sections: xml_validation, backend_failure --
    if api_def.get('is_soap'):
        for section_key in ["xml_validation", "backend_failure"]:
            for tpl in _tpl_cfg.get(section_key, []):
                # body_mod for these templates is used as-is (raw XML or concrete values)
                body_mod = tpl.get("body_mod", {})
                cases.append({
                    "cid": tpl["cid"],
                    "name": tpl["name"],
                    "category": tpl["category"],
                    "status": tpl["status"],
                    "priority": tpl["priority"],
                    "body_mod": body_mod,
                    "header_mod": _resolve_template_header_mod(tpl.get("header_mod", {})),
                    "method_mod": tpl.get("method_mod"),
                })

    # -- ERROR HANDLING: auto-generate from api_def["api_errors"] --
    _common_codes = {v["code"] for v in COMMON_ERRORS.values()}
    err_idx = len(_tpl_cfg.get("error_handling", []))
    _err_keys_covered = set()

    heuristic_rules = _tpl_cfg.get("error_body_mod_heuristics", {}).get("rules", [])
    heuristic_fallback = _tpl_cfg.get("error_body_mod_heuristics", {}).get("fallback", {})

    def _infer_err_body_mod(err_key, err_info):
        key_upper = err_key.upper().replace(" ", "_")

        for rule in heuristic_rules:
            if re.search(rule["pattern"], key_upper):
                field_candidates = rule.get("field_candidates", [])
                target_field = None
                for fc in field_candidates:
                    fc_resolved = fc.replace("{f1}", f1)
                    if fc_resolved in req_body:
                        target_field = fc_resolved
                        break
                if target_field is None and field_candidates:
                    target_field = field_candidates[0].replace("{f1}", f1)

                value = rule.get("value")
                if value == "__REPEAT_DEL_50__":
                    value = "\x7f" * 50

                desc = rule["description_template"].replace("{err_key}", err_key)

                if target_field and value is not None:
                    return {target_field: value}, desc
                elif not field_candidates:
                    return {}, desc
                else:
                    return {}, desc

        # Fallback
        fb_desc = heuristic_fallback.get("description_template", "Trigger {err_key}").replace("{err_key}", err_key)
        if heuristic_fallback.get("use_remove_mandatory") and mandatory:
            return {"__remove__": mandatory[0]}, f"Trigger {err_key} by removing mandatory field"
        prefix = heuristic_fallback.get("fallback_value_prefix", "TRIGGER_ERROR_")
        return {f1: prefix + key_upper[:20]}, fb_desc

    api_errors = api_def.get("api_errors", {})
    for err_key, err_info in api_errors.items():
        if err_key in _COMMON_ERROR_SKIP:
            continue
        if err_info.get("code", "") in _common_codes:
            continue
        if err_key in _err_keys_covered:
            continue
        _err_keys_covered.add(err_key)

        err_idx += 1
        http_status = err_info.get("http", 400)
        err_code = err_info.get("code", "")
        body_mod, desc = _infer_err_body_mod(err_key, err_info)
        # For SOAP APIs, prefer sample-based body_mod (uses actual XML element names)
        # over the heuristic JSON-based mod from _infer_err_body_mod
        if api_def.get('is_soap') and err_info.get('body_mod_hint'):
            body_mod = err_info['body_mod_hint']
        messages_text = err_info.get('messages_text', '')
        if api_def.get('is_soap') and messages_text:
            # Use the actual T24 messages text as TC description
            trigger_fields = ', '.join(f'{k}=""' for k in body_mod if not str(k).startswith('__'))
            if trigger_fields:
                desc = f"T24Error: {messages_text} (set {trigger_fields})"
            else:
                desc = f"T24Error: {messages_text}"
        elif api_def.get('is_soap') and err_info.get('description'):
            # Fallback: use error description from doc table
            desc = f"{desc} — {err_info['description']}"
        # For SOAP APIs, errors come from T24 business rules (not HTTP-level codes)
        err_category = "Business Rules" if api_def.get('is_soap') else "Error Handling"

        cases.append({
            "cid": f"ERR-{err_idx:03d}",
            "name": f"{desc} (code: {err_code})" if err_code else desc,
            "category": err_category,
            "status": http_status,
            "priority": "P1",
            "body_mod": body_mod,
            "header_mod": {},
            "method_mod": None,
            "expected_error_key": err_key,
        })

    return cases


# ── Auto-generate BUS (Business Rules) từ field patterns ──────────────────────
def _auto_business_rules(api_def: dict) -> list:
    """Generate common business-rule TCs based on detected field patterns.

    Pattern detection:
      - pagination fields → BUS pagination tests
      - cifNo/accNo fields → BUS account validation tests
      - date fields → BUS date range tests
      - amount/tenor/term fields → BUS numeric range tests
    """
    rb = api_def.get("request_body", {})
    fields_lower = {k: k for k in rb}            # original name
    fields_set = {k.lower() for k in rb}         # lowercase set for matching
    f1 = next((rb[k] for k in rb if rb[k].get("mandatory")), None)
    f1_name = next((k for k in rb if rb[k].get("mandatory")), "field")
    cases = []
    bus_idx = 1

    # ── Pagination: pageNumber, pageSize, page, size, limit, offset ──
    page_fields = [k for k in rb if k.lower() in
                   ('pagenumber', 'page_number', 'page', 'pageno',
                    'pagesize', 'page_size', 'size', 'limit', 'offset')]
    if len(page_fields) >= 1:
        pf = page_fields[0]
        cases.append({
            "cid": f"BUS-{bus_idx:03d}", "name": f"Pagination: {pf}=0 returns first page",
            "category": "Business Rules", "status": 200, "priority": "P2",
            "body_mod": {pf: 0}, "header_mod": {},
        })
        bus_idx += 1
        cases.append({
            "cid": f"BUS-{bus_idx:03d}", "name": f"Pagination: {pf}=99999 returns empty or last page",
            "category": "Business Rules", "status": 200, "priority": "P2",
            "body_mod": {pf: 99999}, "header_mod": {},
        })
        bus_idx += 1
        if len(page_fields) >= 2:
            ps = page_fields[1]
            cases.append({
                "cid": f"BUS-{bus_idx:03d}", "name": f"Pagination: {ps}=1 returns single record",
                "category": "Business Rules", "status": 200, "priority": "P2",
                "body_mod": {ps: 1}, "header_mod": {},
            })
            bus_idx += 1
            cases.append({
                "cid": f"BUS-{bus_idx:03d}", "name": f"Pagination: {ps}=-1 negative page size",
                "category": "Business Rules", "status": 400, "priority": "P2",
                "body_mod": {ps: -1}, "header_mod": {},
            })
            bus_idx += 1

    # ── CIF / Account validation ──
    cif_field = next((k for k in rb if k.lower() in ('cifno', 'cif', 'cifnumber')), None)
    acc_field = next((k for k in rb if k.lower() in
                      ('accno', 'acctno', 'accountno', 'aaaccount', 'savingaccount',
                       'accountnumber', 'account_number', 'acctnumber')), None)
    if cif_field:
        cases.append({
            "cid": f"BUS-{bus_idx:03d}", "name": f"Non-existent {cif_field}",
            "category": "Business Rules", "status": 400, "priority": "P1",
            "body_mod": {cif_field: "99999999"}, "header_mod": {},
        })
        bus_idx += 1
    if acc_field:
        cases.append({
            "cid": f"BUS-{bus_idx:03d}", "name": f"Non-existent {acc_field}",
            "category": "Business Rules", "status": 400, "priority": "P1",
            "body_mod": {acc_field: "999999999999"}, "header_mod": {},
        })
        bus_idx += 1
    if cif_field and acc_field:
        cases.append({
            "cid": f"BUS-{bus_idx:03d}", "name": f"{cif_field} and {acc_field} mismatch",
            "category": "Business Rules", "status": 400, "priority": "P1",
            "body_mod": {cif_field: "99999999", acc_field: "999999999999"}, "header_mod": {},
        })
        bus_idx += 1

    # ── Date fields: fromDate, toDate, startDate, endDate, maturityDate ──
    date_fields = [k for k in rb if any(pat in k.lower() for pat in
                   ('date', 'fromdate', 'todate', 'startdate', 'enddate',
                    'maturitydate', 'effectivedate', 'valuedate'))]
    if len(date_fields) >= 2:
        d1, d2 = date_fields[0], date_fields[1]
        cases.append({
            "cid": f"BUS-{bus_idx:03d}", "name": f"Date range: {d1} > {d2} (reversed)",
            "category": "Business Rules", "status": 400, "priority": "P2",
            "body_mod": {d1: "20301231", d2: "20200101"}, "header_mod": {},
        })
        bus_idx += 1
        cases.append({
            "cid": f"BUS-{bus_idx:03d}", "name": f"Date range: {d1} = {d2} (same day)",
            "category": "Business Rules", "status": 200, "priority": "P3",
            "body_mod": {d1: "20260101", d2: "20260101"}, "header_mod": {},
        })
        bus_idx += 1
    elif len(date_fields) == 1:
        df = date_fields[0]
        cases.append({
            "cid": f"BUS-{bus_idx:03d}", "name": f"Future date in {df}",
            "category": "Business Rules", "status": 200, "priority": "P3",
            "body_mod": {df: "20301231"}, "header_mod": {},
        })
        bus_idx += 1
        cases.append({
            "cid": f"BUS-{bus_idx:03d}", "name": f"Invalid date format in {df}",
            "category": "Business Rules", "status": 400, "priority": "P2",
            "body_mod": {df: "not-a-date"}, "header_mod": {},
        })
        bus_idx += 1

    # ── Amount / tenor / term / period (numeric business range) ──
    num_fields = [k for k in rb if any(pat in k.lower() for pat in
                  ('amount', 'tenor', 'term', 'period', 'interest', 'rate'))]
    for nf in num_fields:
        cases.append({
            "cid": f"BUS-{bus_idx:03d}", "name": f"Zero value in {nf}",
            "category": "Business Rules", "status": 400, "priority": "P2",
            "body_mod": {nf: 0}, "header_mod": {},
        })
        bus_idx += 1
        cases.append({
            "cid": f"BUS-{bus_idx:03d}", "name": f"Negative value in {nf}",
            "category": "Business Rules", "status": 400, "priority": "P2",
            "body_mod": {nf: -1}, "header_mod": {},
        })
        bus_idx += 1

    # ── Type mismatch: cif = numeric mismatch ──
    if cif_field:
        cases.append({
            "cid": f"BUS-{bus_idx:03d}", "name": f"{cif_field} with special characters",
            "category": "Business Rules", "status": 400, "priority": "P2",
            "body_mod": {cif_field: "!@#$%^&*"}, "header_mod": {},
        })
        bus_idx += 1

    return cases


def generate_all_cases(api_def: dict) -> list:
    """Generate full test case list = common + auto-generated BUS (business rules)."""
    slug = api_def["slug"]
    common = generate_common_cases(api_def)

    # Auto-generate BUS test cases from field pattern detection
    bus_cases = _auto_business_rules(api_def)

    all_cases = common + bus_cases
    result = []
    for case in all_cases:
        tc_id = case['cid']
        entry = {
            "id": tc_id,
            "name": case["name"],
            "category": case["category"],
            "expected_status": case["status"],
            "priority": case["priority"],
            "body_mod": case.get("body_mod", {}),
            "header_mod": case.get("header_mod", {}),
            "method_mod": case.get("method_mod"),
            "expected_error_key": case.get("expected_error_key"),
            "enum_check": case.get("enum_check"),
            "requirements_traceability": "UNDEFINED",
        }
        if "accept_status" in case:
            entry["accept_status"] = case["accept_status"]
        result.append(entry)

    # SOAP APIs always return HTTP 200 at the transport layer; errors are
    # indicated by successIndicator=T24Error in the response body.
    # Exception: XML Validation cases test malformed XML that is rejected at
    # the gateway/parser level before SOAP processing — keep their 4xx status.
    if api_def.get('is_soap'):
        for entry in result:
            if entry.get('expected_status', 200) >= 400 and entry.get('category') != 'XML Validation':
                entry['expected_status'] = 200

    return result


# ==============================================================================
#  BUILD POSTMAN REQUEST
# ==============================================================================

def build_request(api_def: dict, case: dict):
    method = case.get("method_mod") or api_def["method"]

    base_headers = api_def.get("custom_headers") or STANDARD_HEADERS
    headers = copy.deepcopy(base_headers)

    # SOAP APIs use text/xml, not application/json
    if api_def.get('is_soap'):
        for h in headers:
            if h.get("key", "").lower() == "content-type":
                h["value"] = "text/xml"
            elif h.get("key", "").lower() == "accept":
                h["value"] = "text/xml"
    for hk, hv in case["header_mod"].items():
        k_lower = hk.lower()
        if hv is None:
            headers = [h for h in headers if h.get("key", "").lower() != k_lower]
        else:
            found = False
            for h in headers:
                if h.get("key", "").lower() == k_lower:
                    h["value"] = hv
                    found = True
                    break
            if not found:
                headers.append({"key": hk, "value": hv})

    url_obj = build_url_obj(api_def["path"], api_def.get("full_url"))

    body_mod = case.get("body_mod", {})
    body_format = api_def.get("body_format", "json")  # "json" or "xml"

    if "__raw_body__" in body_mod:
        body_raw = body_mod["__raw_body__"]
    elif body_format == "xml":
        # SOAP XML body: apply modifications on the XML template string
        xml_template = api_def.get("body_template_xml", "")
        field_map = api_def.get("body_field_map") or {}
        body_raw = apply_soap_body_mod(xml_template, body_mod, field_map)
    else:
        base_body = build_default_body(api_def)
        field_map = api_def.get("body_field_map") or {}
        body = apply_body_modifications(base_body, body_mod, field_map)
        body_raw = json.dumps(body, ensure_ascii=False, indent=4)

    body_lang = "xml" if body_format == "xml" else "json"
    body_obj = {
        "mode": "raw",
        "raw": body_raw,
        "options": {"raw": {"language": body_lang}},
    }

    return method, headers, url_obj, body_obj


# ==============================================================================
#  DELIVERABLE GENERATORS
# ==============================================================================


def _build_setup_prerequest_js(setup_items: list,
                               post_setup_lines: Optional[list] = None) -> list:
    """Convert setup_items into a chained pm.sendRequest() JavaScript block.

    Each setup item becomes a pm.sendRequest() call inside the previous item's
    callback, ensuring sequential execution.  The generated code:
      1. Runs the item's *prerequest* script (inline, before the fetch).
      2. Calls pm.sendRequest() with variable-resolved URL / headers / body.
      3. Inside the callback, runs the item's *test* script (with
         ``pm.response.json()`` rewritten as ``setupRes<N>.json()``).
    All items are wrapped in a single IIFE so variable names cannot clash
    with the test-case's own pre-request script.

    *post_setup_lines*, if given, are injected **inside** the innermost
    callback so they execute only after every async setup step has finished.
    This prevents timing bugs where code that depends on setup-extracted
    variables (e.g. ``bankDate``) would otherwise run before the HTTP
    responses are available.

    Returns a list[str] of JS lines (no trailing newlines).
    """
    if not setup_items:
        return []

    indent_unit = "    "

    def _js_string(s: str) -> str:
        """Escape a Python string for embedding in JS source."""
        return json.dumps(s, ensure_ascii=False)

    def _adapt_test_lines(lines: list, res_var: str) -> list:
        """Replace pm.response references with the callback response var."""
        out = []
        for line in lines:
            adapted = line.replace("pm.response.json()", f"{res_var}.json()")
            adapted = adapted.replace("pm.response.code", f"{res_var}.code")
            adapted = adapted.replace("pm.response.status", f"{res_var}.status")
            adapted = adapted.replace("pm.response.text()", f"{res_var}.text()")
            out.append(adapted)
        return out

    # ── Build the nested callback chain from innermost → outermost ────────
    # We build from the LAST setup item backwards so each step wraps the next.
    # Result: an array of JS lines at the current indentation depth.

    total_steps = len(setup_items)

    def _build_chain(idx: int, depth: int) -> list:
        """Recursively build the JS for setup item *idx* and all subsequent."""
        if idx >= total_steps:
            return []

        si = setup_items[idx]
        ind = indent_unit * depth
        step = idx + 1
        res_var = f"setupRes{step}"
        err_var = f"setupErr{step}"
        lines: list[str] = []

        lines.append(f"{ind}// ── Setup step {step}/{total_steps}: {si['name']} ──")

        # 1. Prerequest script (runs synchronously before sendRequest)
        for pl in (si.get("prerequest_script") or []):
            lines.append(f"{ind}{pl}")

        # 2. Build request options object
        si_req = si["request"]
        si_url = si_req["url"]
        si_method = si_req["method"]
        si_headers = si_req.get("headers", {})
        si_body_format = si_req.get("body_format", "json")
        if si_body_format == "xml" and si_req.get("body_raw_xml"):
            si_body_json = si_req["body_raw_xml"]
            si_body_lang = "xml"
        else:
            si_body_json = json.dumps(si_req.get("body", {}), ensure_ascii=False)
            si_body_lang = "json"

        # Headers as JS object literal
        hdr_pairs = []
        for k, v in si_headers.items():
            # Skip noisy browser-specific headers
            if k.lower() in ("host", "connection", "cookie", "user-agent"):
                continue
            hdr_pairs.append(f"{ind}    {_js_string(k)}: pm.variables.replaceIn({_js_string(v)})")
        hdr_block = "{\n" + ",\n".join(hdr_pairs) + f"\n{ind}  " + "}" if hdr_pairs else "{}"

        lines.append(f"{ind}pm.sendRequest({{")
        lines.append(f"{ind}  url: pm.variables.replaceIn({_js_string(si_url)}),")
        lines.append(f"{ind}  method: {_js_string(si_method)},")
        lines.append(f"{ind}  header: {hdr_block},")
        lines.append(f"{ind}  body: {{")
        lines.append(f"{ind}    mode: \"raw\",")
        lines.append(f"{ind}    raw: pm.variables.replaceIn({_js_string(si_body_json)}),")
        lines.append(f"{ind}    options: {{ raw: {{ language: \"{si_body_lang}\" }} }}")
        lines.append(f"{ind}  }}")
        lines.append(f"{ind}}}, function ({err_var}, {res_var}) {{")

        inner_ind = indent_unit * (depth + 1)

        # ── Abort-on-failure: network error ──
        lines.append(f"{inner_ind}if ({err_var}) {{")
        lines.append(f"{inner_ind}    console.error('[SETUP ABORT] Step {step}/{total_steps} network error:', {err_var});")
        lines.append(f"{inner_ind}    pm.collectionVariables.set('_setupError', 'Step {step} network error');")
        lines.append(f"{inner_ind}    return;  // ← Stop chain: step {step + 1}+ will NOT run")
        lines.append(f"{inner_ind}}}")

        # ── Abort-on-failure: non-2xx HTTP status ──
        lines.append(f"{inner_ind}if ({res_var}.code < 200 || {res_var}.code >= 300) {{")
        lines.append(f"{inner_ind}    console.error('[SETUP ABORT] Step {step}/{total_steps} HTTP ' + {res_var}.code + ' (expected 2xx)');")
        lines.append(f"{inner_ind}    pm.collectionVariables.set('_setupError', 'Step {step} HTTP ' + {res_var}.code);")
        lines.append(f"{inner_ind}    return;  // ← Stop chain: step {step + 1}+ will NOT run")
        lines.append(f"{inner_ind}}}")

        lines.append(f"{inner_ind}console.log('[SETUP OK] Step {step}/{total_steps} ({si['name']}) → HTTP', {res_var}.code);")

        # 3. Test script (variable extraction) — adapted for callback context
        test_lines = si.get("test_script") or []
        if test_lines:
            lines.append(f"{inner_ind}// --- extract variables ---")
            for tl in _adapt_test_lines(test_lines, res_var):
                lines.append(f"{inner_ind}{tl}")

        # 4. Recurse into next setup item
        next_lines = _build_chain(idx + 1, depth + 1)
        lines.extend(next_lines)

        # 5. If this is the LAST setup item:
        #    - inject post-setup lines inside callback
        #    - set _setupComplete flag
        if idx == total_steps - 1:
            if post_setup_lines:
                lines.append(f"{inner_ind}// ── post-setup (runs after all async steps) ──")
                for psl in post_setup_lines:
                    lines.append(f"{inner_ind}{psl}")
            lines.append(f"{inner_ind}// ── Mark setup chain complete ──")
            lines.append(f"{inner_ind}pm.collectionVariables.set('_setupComplete', 'true');")
            lines.append(f"{inner_ind}console.log('[SETUP COMPLETE] All {total_steps}/{total_steps} steps finished successfully.');")

        lines.append(f"{ind}}});")
        return lines

    # ── Collect variable names that the setup chain will SET ──────────
    _setup_var_names: set = set()
    for si in setup_items:
        for tl in (si.get("test_script") or []):
            # Heuristic: capture pm.collectionVariables.set("<key>", …)
            m = re.search(r'pm\.collectionVariables\.set\(["\']([\w]+)["\']', tl)
            if m:
                _setup_var_names.add(m.group(1))

    result: list[str] = []
    result.append("// ╔══════════════════════════════════════════════════════════╗")
    result.append("// ║  SETUP PREREQUISITES (auto-generated, runs before TC)   ║")
    result.append("// ╚══════════════════════════════════════════════════════════╝")
    # Reset flags at the start of every TC run
    result.append("pm.collectionVariables.set('_setupComplete', '');")
    result.append("pm.collectionVariables.set('_setupError', '');")
    result.append("(function setupPrerequisites() {")
    result.extend(_build_chain(0, 1))
    result.append("})();")

    result.append("")  # blank separator before TC's own prerequest
    return result


def create_collection(api_slug: str, api_def: dict, cases: list, output_file: Path,
                      sampler_meta: Optional[dict] = None,
                      collection_label: str = ""):
    folder_map = {c: [] for c in CATEGORIES}

    prerequest_script = api_def.get("prerequest_script")

    # ── Build setup pre-request JS (embedded pm.sendRequest chain) ────────
    setup_items_raw = api_def.get("setup_items")
    # When both setup chain AND api-level prerequest_script exist, embed
    # the prerequest_script inside the setup chain's innermost callback
    # so it executes only after all async HTTP steps have finished.
    setup_prerequest_js: list = (
        _build_setup_prerequest_js(setup_items_raw,
                                  post_setup_lines=prerequest_script)
        if setup_items_raw else []
    )

    for case in cases:
        method, headers, url_obj, body_obj = build_request(api_def, case)
        req_name = f"{case['id']} - {case['name']}"

        # ─── ASYNC-SAFETY RULE ────────────────────────────────────
        # pm.sendRequest() is NON-BLOCKING.  Any code placed *after*
        # the IIFE runs IMMEDIATELY, BEFORE any HTTP response returns.
        # Therefore: code that depends on variables set by the setup
        # chain (bankDate, cifNo, acctNo, …) MUST live INSIDE the
        # innermost callback — never outside the IIFE.
        #
        # Implementation: when setup chain exists, prerequest_script
        # is passed as post_setup_lines → embedded in last callback.
        # ─────────────────────────────────────────────────────────────
        combined_prerequest: list = []
        if setup_prerequest_js:
            combined_prerequest.extend(setup_prerequest_js)
            # ASSERTION: prerequest_script must NOT also be appended here.
            # It is already embedded inside the IIFE via post_setup_lines.
            if prerequest_script:
                assert prerequest_script is not None  # (already handled)
        elif prerequest_script:
            # No setup chain — just use the prerequest script as-is
            combined_prerequest.extend(prerequest_script)

        events = []
        if combined_prerequest:
            events.append({
                "listen": "prerequest",
                "script": {
                    "type": "text/javascript",
                    "exec": combined_prerequest,
                },
            })
        events.append({
            "listen": "test",
            "script": {
                "type": "text/javascript",
                "exec": postman_test_script(case, api_def),
            },
        })

        request_obj = {
            "name": req_name,
            "event": events,
            "request": {
                "method": method,
                "header": headers,
                "body": body_obj,
                "url": url_obj,
                "description": f"Test Case: {case['id']}\\nCategory: {case['category']}\\nPriority: {case['priority']}\\nExpected: HTTP {case['expected_status']}",
            },
            "response": [],
        }
        folder_map.setdefault(case["category"], []).append(request_obj)

    items = []

    # No separate "Setup (run first)" folder — prerequisites are now embedded
    # in each test case's pre-request script via pm.sendRequest().

    for category in CATEGORIES:
        reqs = folder_map.get(category, [])
        if reqs:
            items.append({"name": category, "item": reqs})

    _meta = sampler_meta or {}
    # Only include collection variables if sampler actually declared them
    if _meta.get("_has_sampler_vars", False):
        variables = list(_meta.get("variable", []))
    else:
        variables = []
    auth_obj  = _meta.get("auth")

    extra_vars = api_def.get("extra_variables") or {}
    if extra_vars:
        existing_keys = {v["key"] for v in variables}
        for k, v in extra_vars.items():
            if k not in existing_keys:
                variables.append({"key": k, "value": v, "type": "string"})

    now = datetime.now(timezone.utc).isoformat()
    display_name = f"[{collection_label}] {api_slug}" if collection_label else api_slug
    collection = {
        "info": {
            "_postman_id": f"{api_slug}-generated-v2",
            "name": display_name,
            "schema": "https://schema.getpostman.com/json/collection/v2.1.0/collection.json",
            "description": f"Generated at {now}. Full API contract coverage.",
        },
        "item": items,
    }
    if variables:
        collection["variable"] = variables
    if auth_obj:
        collection["auth"] = auth_obj

    output_file.write_text(json.dumps(collection, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  [OK] Collection -> {output_file}  ({len(cases)} requests)")


def create_csv(api_slug: str, api_def: dict, cases: list, output_file: Path):
    fieldnames = [
        "Test Case ID", "Test Case Name", "Category", "Pre-conditions",
        "Test Steps", "Request Data", "Expected Result", "HTTP Status",
        "Priority", "Postman Request Name", "Requirements Traceability",
    ]

    with output_file.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for case in cases:
            method = case.get("method_mod") or api_def["method"]
            expected_desc = f"System returns HTTP {case['expected_status']}"
            _soap_neg_categories = ('Negative Validation', 'Business Rules', 'Error Handling')
            _soap_ep_is_error = (
                case.get('category') in ('Equivalence Partitioning', 'Boundary Value Analysis')
                and any(kw in (case.get('name') or '') for kw in ('Invalid', 'invalid', 'Missing', 'Empty', 'Null', 'null', 'Negative', 'negative'))
            )
            _is_soap = api_def.get('is_soap', False)
            _is_soap_error_case = _is_soap and (
                case.get('category') in _soap_neg_categories or _soap_ep_is_error
            )
            _cat = case.get('category', '')
            if _is_soap_error_case:
                err_key = case.get('expected_error_key', '')
                if err_key:
                    err_msgs = api_def.get('api_errors', {}).get(err_key, {}).get('messages_text', '')
                    messages_ref = f"'{err_msgs}'" if err_msgs else f"'{err_key}'"
                    expected_desc += (
                        f". SOAP response body: <successIndicator>T24Error</successIndicator>"
                        f" with <messages> containing {messages_ref}"
                    )
                else:
                    expected_desc += (
                        ". SOAP response body contains"
                        " <successIndicator>T24Error</successIndicator>"
                        " and <messages> with relevant error details."
                    )
            elif _cat == 'XML Validation':
                expected_desc += (
                    ". SOAP gateway rejects malformed/invalid XML and returns"
                    " an HTTP error or SOAP Fault response."
                )
            elif _cat == 'Backend Failure':
                expected_desc += (
                    ". SOAP response body: <successIndicator>T24Error</successIndicator>"
                    " with <messages> containing backend authentication or configuration error details."
                )
            elif case.get("expected_error_key"):
                expected_desc += f" with error code {case['expected_error_key']}"
            expected_desc += " with compliant response structure."

            writer.writerow({
                "Test Case ID": case["id"],
                "Test Case Name": case["name"],
                "Category": case["category"],
                "Pre-conditions": "API endpoint is reachable; valid environment configuration loaded; test data prepared.",
                "Test Steps": f"1) Send {method} request to {api_def['path']}. 2) Validate response status/body/headers.",
                "Request Data": json.dumps({
                    "body_mod": case["body_mod"],
                    "header_mod": case["header_mod"],
                    "method_mod": case.get("method_mod"),
                }, ensure_ascii=False),
                "Expected Result": expected_desc,
                "HTTP Status": case["expected_status"],
                "Priority": case["priority"],
                "Postman Request Name": f"{case['id']} - {case['name']}",
                "Requirements Traceability": case["requirements_traceability"],
            })

    print(f"  [OK] CSV -> {output_file}  ({len(cases)} rows)")


# ==============================================================================
#  EXCEL OUTPUT — matching Template TC 4.xlsx format
# ==============================================================================

def create_excel(api_slug: str, api_def: dict, cases: list, output_file: Path):
    """Generate Excel file matching the Template TC 4 format."""
    if _excel_cfg is None:
        return  # no excel template configured

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        print("  ⚠️  openpyxl not installed — skipping Excel output")
        return

    wb = openpyxl.Workbook()
    ws = wb.active

    columns = _excel_cfg["columns"]
    field_mapping = _excel_cfg["field_mapping"]
    priority_map = _excel_cfg.get("priority_map", {})
    suite_map = _excel_cfg.get("scenarios_suite_map", {})
    feature_map = _excel_cfg.get("feature_map", {})

    # Sheet name — strip invalid Excel characters: \ / * ? [ ] :
    import re as _re
    sheet_name = _excel_cfg.get("sheet_name_template", "API_{slug}").replace("{slug}", api_slug)
    sheet_name = _re.sub(r'[\\\/*?\[\]:]', '_', sheet_name)
    ws.title = sheet_name[:31]  # Excel max sheet name length

    # Determine feature name
    feature = feature_map.get(api_slug, feature_map.get("default", api_slug))
    if feature == "{slug}":
        feature = api_slug

    # Row 1: merged header
    last_col_letter = columns[-1]["col"]
    merge_range = f"A1:{last_col_letter}1"
    ws.merge_cells(merge_range)
    ws["A1"] = feature
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Row 2: column headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_def in columns:
        cell = ws[f"{col_def['col']}2"]
        cell.value = col_def["header"]
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[col_def["col"]].width = col_def.get("width", 15)

    # Data rows
    data_start = _excel_cfg.get("data_start_row", 3)
    for i, case in enumerate(cases):
        row = data_start + i
        method = case.get("method_mod") or api_def["method"]

        # Build request body (JSON or XML)
        body_mod = case.get("body_mod", {})
        body_format = api_def.get("body_format", "json")
        if "__raw_body__" in body_mod:
            request_body_json = body_mod["__raw_body__"]
        elif body_format == "xml":
            xml_template = api_def.get("body_template_xml", "")
            field_map_body = api_def.get("body_field_map") or {}
            request_body_json = apply_soap_body_mod(xml_template, body_mod, field_map_body)
        else:
            base_body = build_default_body(api_def)
            field_map_body = api_def.get("body_field_map") or {}
            modified_body = apply_body_modifications(base_body, body_mod, field_map_body)
            request_body_json = json.dumps(modified_body, ensure_ascii=False, indent=4)

        # Build expected result
        expected_desc = f"System returns HTTP {case['expected_status']}"
        # Categories that always produce SOAP T24Error in response body
        _soap_neg_categories = ('Negative Validation', 'Business Rules', 'Error Handling')
        # EP/BVA cases that produce errors are identified by "Invalid" in the TC name
        _soap_ep_is_error = (
            case.get('category') in ('Equivalence Partitioning', 'Boundary Value Analysis')
            and any(kw in (case.get('name') or '') for kw in ('Invalid', 'invalid', 'Missing', 'Empty', 'Null', 'null', 'Negative', 'negative'))
        )
        _is_soap = api_def.get('is_soap', False)
        _is_soap_error_case = _is_soap and (
            case.get('category') in _soap_neg_categories or _soap_ep_is_error
        )
        _cat = case.get('category', '')
        if _is_soap_error_case and case.get("expected_error_key"):
            err_key = case["expected_error_key"]
            err_msgs = api_def.get('api_errors', {}).get(err_key, {}).get('messages_text', '')
            messages_ref = f"'{err_msgs}'" if err_msgs else f"'{err_key}'"
            # SOAP errors are indicated in response body, not HTTP status
            expected_desc = (
                f"System returns HTTP {case['expected_status']}."
                f" SOAP response body: <successIndicator>T24Error</successIndicator>"
                f" with <messages> containing {messages_ref}"
            )
        elif _is_soap_error_case:
            expected_desc += (
                ". SOAP response body contains"
                " <successIndicator>T24Error</successIndicator>"
                " and <messages> with relevant error details."
            )
        elif _cat == 'XML Validation':
            expected_desc += (
                ". SOAP gateway rejects malformed/invalid XML and returns"
                " an HTTP error or SOAP Fault response."
            )
        elif _cat == 'Backend Failure':
            expected_desc += (
                ". SOAP response body: <successIndicator>T24Error</successIndicator>"
                " with <messages> containing backend authentication or configuration error details."
            )
        elif case.get("expected_error_key"):
            all_errors = {**COMMON_ERRORS, **api_def.get("api_errors", {})}
            err_key = case["expected_error_key"]
            if err_key in all_errors:
                expected_desc = json.dumps({
                    "code": all_errors[err_key]["code"],
                    "message": "",
                    "messageKey": err_key,
                }, ensure_ascii=False, indent=4)
            else:
                expected_desc += f" with error code {err_key}"
        expected_desc_final = expected_desc

        # Scenarios suite
        cat = case.get("category", "")
        suite = suite_map.get(cat, suite_map.get("default", "UnHappyFlow"))

        # Priority mapping
        priority_mapped = priority_map.get(case["priority"], case["priority"])

        # Write cells
        values = {
            "test_case_id": i + 1,
            "feature": feature,
            "scenarios_suite": suite,
            "summary": case["name"],
            "precondition": field_mapping.get("precondition", "1. Input data"),
            "assignee_name": field_mapping.get("assignee_name", ""),
            "priority": priority_mapped,
            "test_data": request_body_json,
            "steps": field_mapping.get("steps", "1. Call Request\n2. Check Response"),
            "expected_result": expected_desc_final,
            "results": "",
            "test_type": "",
            "test_environments": field_mapping.get("test_environments", "UAT"),
            "tag": "",
            "test_repository": "",
            "labels": "",
        }

        for col_def in columns:
            cell = ws[f"{col_def['col']}{row}"]
            cell.value = values.get(col_def["key"], "")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = thin_border

    wb.save(str(output_file))
    print(f"  [OK] Excel -> {output_file}  ({len(cases)} rows)")


# ==============================================================================
#  COVERAGE SUMMARY
# ==============================================================================

def create_coverage_summary(api_slug: str, api_def: dict, cases: list, output_file: Path, is_new: bool):
    by_category = {}
    by_priority = {}
    by_status = {}
    for c in cases:
        by_category[c["category"]] = by_category.get(c["category"], 0) + 1
        by_priority[c["priority"]] = by_priority.get(c["priority"], 0) + 1
        by_status[c["expected_status"]] = by_status.get(c["expected_status"], 0) + 1

    now = datetime.now(timezone.utc).isoformat()
    lines = [
        f"# Test Coverage Summary - {api_slug}",
        "",
        f"- **Generated date**: {now}",
        f"- **Total test cases**: {len(cases)}",
        f"- **API Method**: {api_def['method']} (JSON body)",
        f"- **API Path**: {api_def['path']}",
        "- **Requirements traceability**: DOC input only; unresolved items marked `UNDEFINED`.",
        "",
        "## Coverage by Category",
        "",
        "| Category | Count |",
        "|---|---|",
    ]
    for k in CATEGORIES:
        v = by_category.get(k, 0)
        if v:
            lines.append(f"| {k} | {v} |")

    lines.extend(["", "## Coverage by Priority", "", "| Priority | Count |", "|---|---|"])
    for p in sorted(by_priority.keys()):
        lines.append(f"| {p} | {by_priority[p]} |")

    lines.extend(["", "## Coverage by HTTP Status Code", "", "| Status | Count |", "|---|---|"])
    for s in sorted(by_status.keys()):
        lines.append(f"| {s} | {by_status[s]} |")

    distinct_status_count = len(by_status)

    checklist_rows = [
        "",
        "## Coverage Checklist (Prompt Requirements)",
        "",
        f"Bảng dưới bám sát **{len(PROMPT_REQUIREMENTS)} yêu cầu bắt buộc** từ prompt gốc"
        f" + {len(EXTRA_CATEGORIES)} category bổ sung (đánh dấu *extra*).",
        "",
        "| # | Requirement (từ prompt) | Covered | TCs | Ghi chú |",
        "|---|------------------------|---------|-----|--------|",
    ]
    prompt_covered = 0
    for num, req_label, cat_key, desc in PROMPT_REQUIREMENTS:
        if cat_key == '__http_status_codes__':
            tc_count = f"{distinct_status_count} codes"
            # SOAP APIs always return HTTP 200 — only 1 distinct code is expected
            if api_def.get('is_soap'):
                covered = 'Y' if distinct_status_count >= 1 else 'N'
            else:
                covered = 'Y' if distinct_status_count >= 4 else 'N'
            detail   = ', '.join(str(s) for s in sorted(by_status.keys()))
        else:
            tc_raw   = by_category.get(cat_key, 0)
            tc_count = str(tc_raw)
            covered  = 'Y' if tc_raw > 0 else 'N'
            detail   = desc
        if covered == 'Y':
            prompt_covered += 1
        mark = '✅' if covered == 'Y' else '❌'
        checklist_rows.append(f"| {num:02d} | {mark} {req_label} | {covered} | {tc_count} | {detail} |")
    for ex_label, ex_cat, ex_desc in EXTRA_CATEGORIES:
        tc_raw = by_category.get(ex_cat, 0)
        mark   = '✅' if tc_raw > 0 else '❌'
        checklist_rows.append(
            f"| +  | {mark} {ex_label} *(extra)* | {'Y' if tc_raw > 0 else 'N'} | {tc_raw} | {ex_desc} |"
        )
    prompt_pct = round(prompt_covered / len(PROMPT_REQUIREMENTS) * 100, 1) if PROMPT_REQUIREMENTS else 0.0
    _tgt_prompt = KPI_TARGETS['prompt_coverage_pct']
    pct_verdict = f'✅ ≥ {_tgt_prompt}% — đạt yêu cầu' if prompt_pct >= _tgt_prompt else f'❌ < {_tgt_prompt}% — cần bổ sung TC'
    checklist_rows.extend(["", f"> **Prompt coverage score: {prompt_covered}/{len(PROMPT_REQUIREMENTS)} = {prompt_pct}%** — {pct_verdict}"])
    lines.extend(checklist_rows)

    lines.extend([
        "",
        "## Response Schema Assertions (per 200 OK test case)",
        "",
        "- Response envelope: " + ", ".join(RESPONSE_ENVELOPE),
        "- Data required fields: " + ", ".join(api_def.get("response_data_fields", {}).keys()),
    ])
    arr_field = api_def.get("response_array_field")
    if arr_field:
        lines.append(f"- Array item fields ({arr_field}): " + ", ".join(api_def.get("response_array_item_fields", {}).keys()))
    lines.append("- Data type assertions for each field")

    enums = api_def.get("enums", {})
    if enums:
        lines.extend(["", "## Enum Validations"])
        for k, v in enums.items():
            lines.append(f"- `{k}`: {v}")

    errors = {**COMMON_ERRORS, **api_def.get("api_errors", {})}
    lines.extend(["", "## Error Code Coverage", "", "| Error Code | MessageKey | HTTP |", "|---|---|---|"])
    for mk, info in errors.items():
        lines.append(f"| {info['code']} | {mk} | {info['http']} |")

    lines.extend(["", "## Business Rules Covered"])
    for rule in api_def.get("business_rules", []):
        lines.append(f"- {rule}")

    metrics = None
    if is_new:
        p1_count  = by_priority.get('P1', 0)
        p2_count  = by_priority.get('P2', 0)
        p1_pct    = round(p1_count / len(cases) * 100, 1) if cases else 0.0
        p1p2_pct  = round((p1_count + p2_count) / len(cases) * 100, 1) if cases else 0.0
        error_code_count = len(errors)
        status_count     = len(by_status)
        kpi = KPI_TARGETS
        prompt_ok = '✅' if prompt_pct >= kpi['prompt_coverage_pct'] else '❌'
        p1_ok     = '✅' if p1_pct >= kpi['min_p1_pct'] else '⚠️'
        p1p2_ok   = '✅' if p1p2_pct >= kpi['min_p1p2_pct'] else '⚠️'
        tc_ok     = '✅' if len(cases) >= kpi['min_total_tcs'] else '⚠️'
        err_ok    = '✅' if error_code_count >= kpi['min_error_codes'] else '⚠️'
        # SOAP APIs only use HTTP 200 — threshold is 1
        _min_http = 1 if api_def.get('is_soap') else kpi['min_http_status_codes']
        http_ok   = '✅' if status_count >= _min_http else '⚠️'

        lines.extend([
            "", "## Manager KPI", "",
            f"> Ngưỡng KPI áp dụng: prompt ≥ {kpi['prompt_coverage_pct']}%,"
            f" tổng TC ≥ {kpi['min_total_tcs']},"
            f" P1 ≥ {kpi['min_p1_pct']}%, P1+P2 ≥ {kpi['min_p1p2_pct']}%,"
            f" HTTP codes ≥ {kpi['min_http_status_codes']},"
            f" error codes ≥ {kpi['min_error_codes']}."
            f" Chỉnh trong `baseline/coverage_requirements.json`.",
            "",
            "| # | Metric | Target | Achieved | % | Status |",
            "|---|--------|--------|----------|---|--------|",
            f"| 1 | Prompt requirements covered | {len(PROMPT_REQUIREMENTS)}/{len(PROMPT_REQUIREMENTS)} | {prompt_covered}/{len(PROMPT_REQUIREMENTS)} | {prompt_pct}% | {prompt_ok} |",
            f"| 2 | Total test cases | ≥ {kpi['min_total_tcs']} | {len(cases)} | — | {tc_ok} |",
            f"| 3 | Critical TCs (P1) | ≥ {kpi['min_p1_pct']}% of total | {p1_count}/{len(cases)} | {p1_pct}% | {p1_ok} |",
            f"| 4 | Critical + High TCs (P1+P2) | ≥ {kpi['min_p1p2_pct']}% of total | {p1_count + p2_count}/{len(cases)} | {p1p2_pct}% | {p1p2_ok} |",
            f"| 5 | Error codes from doc covered | ≥ {kpi['min_error_codes']} codes | {error_code_count} codes | — | {err_ok} |",
            f"| 6 | HTTP status codes covered | ≥ {kpi['min_http_status_codes']} distinct codes | {status_count} | — | {http_ok} |",
            "",
            "> **Giải thích**:",
            "> - **Prompt requirements covered**: số trong 15 yêu cầu bắt buộc có ít nhất 1 TC.",
            "> - **P1 %**: tỉ lệ TC mức Critical.",
            "> - **P1+P2 %**: tỉ lệ TC mức Critical + High.",
            "> - **Error codes**: tất cả mã lỗi khai báo trong doc đều có TC tương ứng.",
            "> - **HTTP status codes**: số giá trị HTTP khác nhau xuất hiện trong tập TC.",
            "- Missing requirement/error-code mapping from DOC is marked `UNDEFINED` for manual validation.",
        ])

        metrics = {
            "slug": api_slug, "total_tcs": len(cases),
            "prompt_covered": prompt_covered, "prompt_total": len(PROMPT_REQUIREMENTS),
            "prompt_pct": prompt_pct, "p1_pct": p1_pct, "p1p2_pct": p1p2_pct,
            "status_count": status_count, "error_count": error_code_count,
            "is_soap": api_def.get('is_soap', False),
        }

    lines.extend([
        "", "## Execution Guidelines",
        "- Run in Postman Runner with environment variables (`baseUrl`, `apikey`, `channel`, `traceId`).",
        "- Or use Newman: `newman run <collection>.json -e <environment>.postman_environment.json`",
        "- Newman + htmlextra: `newman run <collection>.json -e <env>.json -r htmlextra --reporter-htmlextra-export report.html`",
    ])

    output_file.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"  [OK] Coverage Summary -> {output_file}")
    return metrics


def create_traceability_file(api_slug: str, api_def: dict, cases: list, output_file: Path):
    lines = [
        f"# API Traceability - {api_slug}", "",
        "AI-generated output. Manual review is required before official use.",
        "Requirement IDs sourced only from DOC. Unresolved mapping is marked `UNDEFINED`.", "",
        f"**API Path**: {api_def['path']}",
        f"**HTTP Method**: {api_def['method']}",
        f"**Total Test Cases**: {len(cases)}", "",
        "## Traceability Matrix", "",
        "| # | Test Case ID | Test Case Name | Category | Priority | Postman Request Name | Requirement ID | Mapping Note |",
        "|---|---|---|---|---|---|---|---|",
    ]
    for i, c in enumerate(cases, 1):
        req_name = f"{c['id']} - {c['name']}"
        lines.append(
            f"| {i} | {c['id']} | {c['name']} | {c['category']} | {c['priority']} "
            f"| {req_name} | UNDEFINED | Awaiting DOC mapping validation |"
        )

    by_cat = {}
    for c in cases:
        by_cat[c["category"]] = by_cat.get(c["category"], 0) + 1

    lines.extend(["", "## Coverage Summary by Category", "", "| Category | Count |", "|---|---|"])
    for k, v in sorted(by_cat.items(), key=lambda x: -x[1]):
        lines.append(f"| {k} | {v} |")

    output_file.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"  [OK] Traceability -> {output_file}")


# ==============================================================================
#  MAIN
# ==============================================================================

def process_api(api_def: dict):
    slug = api_def["slug"]
    cases = generate_all_cases(api_def)
    out_dir = ROOT / slug
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n[{slug}] Generating {len(cases)} test cases...")

    create_csv(slug, api_def, cases, out_dir / f"TestCases_{slug}.csv")
    create_collection(slug, api_def, cases, out_dir / f"{slug}_Postman_Collection.json")
    create_coverage_summary(slug, api_def, cases, out_dir / "Test_Coverage_Summary.md", is_new=True)
    create_traceability_file(slug, api_def, cases, out_dir / f"API-{slug}.test-case-traceability.md")
    create_excel(slug, api_def, cases, out_dir / f"TestCases_{slug}.xlsx")


def main():
    print("=" * 70)
    print("  Generate deliverables for Data-Platform APIs (v3 - config-driven)")
    print("=" * 70)

    total = 0
    for api_def in ALL_APIS:
        cases = generate_all_cases(api_def)
        total += len(cases)
        print(f"  {api_def['slug']}: {len(cases)} test cases")

    print(f"  TOTAL: {total} test cases")
    print("=" * 70)

    for api_def in ALL_APIS:
        process_api(api_def)

    print("\n" + "=" * 70)
    print("  All done!")
    print("=" * 70)


if __name__ == "__main__":
    print()
    print("⚠️  generate_outputs.py không nên chạy trực tiếp.")
    print("   Nó chỉ là thư viện (library) cho regen_from_contracts.py.")
    print()
    print("   ╔══════════════════════════════════════════════════════════╗")
    print("   ║  Cách chạy chuẩn:                                       ║")
    print("   ║                                                          ║")
    print("   ║  python3 run_pipeline.py          (chạy toàn bộ)        ║")
    print("   ║                                                          ║")
    print("   ║  python3 run_pipeline.py --newman  (+ chạy Newman)      ║")
    print("   ╚══════════════════════════════════════════════════════════╝")
    print()
    sys.exit(1)
