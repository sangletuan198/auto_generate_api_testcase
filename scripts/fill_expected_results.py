#!/usr/bin/env python3
"""
fill_expected_results.py

Sau khi Newman chạy xong:
  - Excel cột "Results"        ← điền actual response từ newman-report.json
  - Excel cột "Expected Result" ← generate lại từ Postman Collection (HTTP status
                                  + error code nếu có), KHÔNG ghi đè bằng response
  - CSV cột "Expected Result"   ← giữ nguyên text generate (không thay đổi)

Usage:
    python3 scripts/fill_expected_results.py \\
        --bundle 20260309-113354 \\
        --target corrected \\
        --api transfer247Payment
"""

import argparse
import csv
import json
import re
import sys
from pathlib import Path
from typing import Optional

ROOT_DIR = Path(__file__).parent.parent

# Fields that change every run → strip from Results display
_EPHEMERAL_FIELDS = {"timeStamp", "timestamp", "traceId", "id", "correlationId"}


# ── Response extraction from Newman ─────────────────────────────────────────

def _decode_stream(stream_obj) -> str:
    if not stream_obj:
        return ""
    data = stream_obj.get("data", [])
    if isinstance(data, list):
        return bytes(data).decode("utf-8", errors="replace")
    if isinstance(data, str):
        return data
    return ""


def extract_responses(newman_json_path: Path) -> dict:
    """
    Build map: tc_name → {short_id, status, body_json, body_text}
    Last occurrence wins (handles pre/post-script duplicate entries).
    """
    raw = json.loads(newman_json_path.read_text(encoding="utf-8"))
    executions = raw.get("run", {}).get("executions", [])

    result = {}

    for ex in executions:
        item_name = ex.get("item", {}).get("name", "").strip()
        if not item_name:
            continue

        if " - " in item_name:
            short_id, tc_name = item_name.split(" - ", 1)
            short_id = short_id.strip()
            tc_name = tc_name.strip()
        else:
            short_id = tc_name = item_name

        resp = ex.get("response")
        if not resp:
            continue

        status = resp.get("code", 0)
        body_text = _decode_stream(resp.get("stream"))

        body_json = None
        try:
            body_json = json.loads(body_text)
        except Exception:
            pass

        result[tc_name] = {
            "short_id": short_id,
            "status": status,
            "body_text": body_text,
            "body_json": body_json,
        }

    return result


# ── Formatting ───────────────────────────────────────────────────────────────

def _strip_ephemeral(obj):
    if not isinstance(obj, dict):
        return obj
    return {k: v for k, v in obj.items() if k not in _EPHEMERAL_FIELDS}


def format_actual_result(resp: dict) -> str:
    """Format actual response for the Results column (strip ephemeral fields)."""
    body_json = resp.get("body_json")
    if body_json is not None:
        clean = _strip_ephemeral(body_json)
        return json.dumps(clean, ensure_ascii=False, indent=2)
    body_text = resp.get("body_text", "").strip()
    if body_text:
        return body_text[:1000]
    return "HTTP {} (empty body)".format(resp.get("status", "?"))


def _build_expected_from_collection(collection_path: Path) -> dict:
    """
    Read Postman collection and build map:
        tc_name → expected_result_text

    The collection request name is like:
        "TRANSFER247PAYMENT-POS-001 - Valid request with all parameters"

    Expected result text mirrors generate_outputs.py logic:
        Positive (2xx): "System returns HTTP <status> with compliant response structure."
        Error test:     "System returns HTTP <status> with error code <KEY> with compliant response structure."
    """
    col = json.loads(collection_path.read_text(encoding="utf-8"))

    result = {}

    def _walk(items):
        for item in items:
            if "item" in item:
                _walk(item["item"])
                continue
            name = item.get("name", "")
            if " - " in name:
                _, tc_name = name.split(" - ", 1)
                tc_name = tc_name.strip()
            else:
                tc_name = name.strip()

            # Find expected status from test script assertions like
            # pm.response.to.have.status(200)  or  pm.expect(pm.response.code).to.equal(404)
            events = item.get("event", [])
            status_code = None
            error_key = None
            for ev in events:
                if ev.get("listen") != "test":
                    continue
                src = " ".join(ev.get("script", {}).get("exec", []))
                m = re.search(r'have\.status\((\d+)\)', src)
                if not m:
                    m = re.search(r'\.equal\((\d{3})\)', src)
                if m:
                    status_code = int(m.group(1))
                # Error key pattern: 'ACCOUNT_NOT_FOUND'
                ek = re.search(r"'([A-Z_]{3,})'\s*\)", src)
                if ek and not re.match(r'^(HTTP|GET|POST|PUT|PATCH|DELETE|VND|JSON)$', ek.group(1)):
                    error_key = ek.group(1)

            if status_code is None:
                status_code = "2xx"

            if error_key:
                desc = "System returns HTTP {} with error code {} with compliant response structure.".format(
                    status_code, error_key
                )
            else:
                desc = "System returns HTTP {} with compliant response structure.".format(status_code)

            result[tc_name] = desc

    _walk(col.get("item", []))
    return result


# ── Matching ─────────────────────────────────────────────────────────────────

def _find_response(tc_id: str, tc_name: str, responses: dict) -> Optional[dict]:
    if tc_name in responses:
        return responses[tc_name]
    for _name, resp in responses.items():
        s_id = resp.get("short_id", "")
        if s_id and (tc_id.endswith("-" + s_id) or tc_id == s_id):
            return resp
    return None


# ── CSV update ───────────────────────────────────────────────────────────────

def update_csv(csv_path: Path, responses: dict, expected_map: dict) -> tuple:
    """
    CSV: only restore Expected Result from generated text.
    Returns (expected_updated, actual_updated).
    """
    rows = []
    fieldnames = []

    with csv_path.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        fieldnames = list(reader.fieldnames or [])
        rows = list(reader)

    exp_col = "Expected Result" in fieldnames
    exp_count = 0

    for row in rows:
        tc_id = row.get("Test Case ID", "").strip()
        tc_name = row.get("Test Case Name", "").strip()

        # Restore Expected Result from collection-derived text
        if exp_col and tc_name in expected_map:
            row["Expected Result"] = expected_map[tc_name]
            exp_count += 1

    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    return exp_count, 0


# ── Excel update ─────────────────────────────────────────────────────────────

def update_excel(xlsx_path: Path, responses: dict, expected_map: dict) -> tuple:
    """
    Excel:
      - col "Expected Result" ← text from generated collection (not raw response)
      - col "Results"         ← actual response from Newman
    Returns (expected_updated, actual_updated).
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment
    except ImportError:
        print("    \u26a0\ufe0f  openpyxl not installed \u2014 skipping Excel update")
        return 0, 0

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # Find header row
    header_row_idx = None
    col_map = {}
    for r in range(1, 6):
        row_vals = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v:
                row_vals[str(v)] = c
        if "Expected Result" in row_vals or "Summary" in row_vals:
            header_row_idx = r
            col_map = row_vals
            break

    if header_row_idx is None:
        print("    \u26a0\ufe0f  Header row not found in {} \u2014 skipping".format(xlsx_path.name))
        return 0, 0

    # Map column names → indices
    expected_col = None
    results_col = None
    id_col = None
    summary_col = None

    for header, col in col_map.items():
        hl = header.lower()
        if "expected result" in hl:
            expected_col = col
        elif hl == "results" or hl == "k\u1ebft qu\u1ea3" or "actual" in hl:
            results_col = col
        elif "test case id" in hl or "tc id" in hl:
            id_col = col
        elif "summary" in hl:
            summary_col = col

    exp_count = 0
    act_count = 0

    wrap_top = Alignment(wrap_text=True, vertical="top")

    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        tc_id = str(ws.cell(row=row_idx, column=id_col).value or "").strip() if id_col else ""
        tc_name = str(ws.cell(row=row_idx, column=summary_col).value or "").strip() if summary_col else ""
        if not tc_id and not tc_name:
            continue

        # Expected Result ← generated text
        if expected_col and tc_name in expected_map:
            cell = ws.cell(row=row_idx, column=expected_col)
            cell.value = expected_map[tc_name]
            cell.alignment = wrap_top
            exp_count += 1

        # Results ← actual response
        if results_col:
            resp = _find_response(tc_id, tc_name, responses)
            if resp:
                cell = ws.cell(row=row_idx, column=results_col)
                cell.value = format_actual_result(resp)
                cell.alignment = wrap_top
                act_count += 1

    wb.save(xlsx_path)
    return exp_count, act_count


# ── Entry point ──────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(
        description="Fill Expected Result (generated) and Results (actual) columns"
    )
    ap.add_argument("--bundle", required=True)
    ap.add_argument("--target", required=True)
    ap.add_argument("--api",    required=True)
    args = ap.parse_args()

    bundle_dir = ROOT_DIR / "output" / "bundles" / args.bundle / args.target / args.api
    api_dir    = ROOT_DIR / "output" / args.target / args.api

    newman_json = bundle_dir / "newman-report.json"
    if not newman_json.exists():
        print("  \u274c  Newman report not found: {}".format(newman_json), file=sys.stderr)
        sys.exit(1)

    # Find Postman collection in api_dir
    collection_path = None
    for p in sorted(api_dir.glob("*_Postman_Collection*.json")):
        collection_path = p
        break
    if not collection_path:
        print("  \u26a0\ufe0f  No Postman collection found in {} \u2014 expected text will be empty".format(api_dir))

    print("  [FILL] Parsing Newman responses from bundle {} ...".format(args.bundle))
    responses = extract_responses(newman_json)
    print("  [FILL] {} unique TCs from Newman report".format(len(responses)))

    expected_map = {}
    if collection_path:
        expected_map = _build_expected_from_collection(collection_path)
        print("  [FILL] {} expected-result texts built from collection".format(len(expected_map)))

    # Update CSV
    for csv_path in sorted(api_dir.glob("TestCases_*.csv")):
        exp_n, _ = update_csv(csv_path, responses, expected_map)
        print("  [FILL] \u2705  CSV: {} Expected Result rows restored in {}".format(exp_n, csv_path.name))

    # Update Excel
    for xlsx_path in sorted(api_dir.glob("TestCases_*.xlsx")):
        exp_n, act_n = update_excel(xlsx_path, responses, expected_map)
        print("  [FILL] \u2705  Excel: {} Expected Result + {} Results rows updated in {}".format(
            exp_n, act_n, xlsx_path.name))


if __name__ == "__main__":
    main()
